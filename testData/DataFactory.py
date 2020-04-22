import openpyxl


class DataFactory:
    # cílem je získat tuto strukturu dynamicky z excelu
    """test_homepage_data_old = [{'firstname': 'Jakub', 'email': 'j@j.cz', 'gender':'Male'},
                            {'firstname': 'Tereza', 'email': 't@t.cz', 'gender': 'Female'}]
    """
    source_path = "C:\\Users\\jawlc\\Disk Google\\testingPython\\SeleniumFramework\\testData\\datasource.xlsx"

    @staticmethod
    def get_homepage_data():
        return DataFactory.collect_data_from_xls()

    @staticmethod
    def collect_data_from_xls():
        book = openpyxl.load_workbook(DataFactory.source_path)
        sheet = book.active
        test_homepage_data = []
        for r in range(2, sheet.max_row + 1):
            data = {}
            for c in range(2, sheet.max_column + 1):
                data[sheet.cell(row=1, column=c).value] = sheet.cell(row=r, column=c).value
            test_homepage_data.append(data)
        return test_homepage_data
